import logging
import openpyxl as px
import pandas as pd

from python.services.functions.mounted_directory import construct_file_path

PROHIBITED_WORDS = [
    'REGISTERED SHARES', ' O N', 'O N', 'INHABER AKTIEN', 'NAMENS AKTIEN', 'REG SHARES CLASS',
    'REG SHARES', 'INHABER STAMMAKTIEN',
    'REG ', 'REG S', 'NAMENSAKTIEN', 'NAM AKT', 'NAMN AKTIER', 'NAVNE AKSJER NK', 'NAVNE AKTIER',
    ' A ', ' LS ', ' DL ', 'DK 1', 'ADR 1', 'ADRS 1', 'DL 10',
    ' 01 ', ' 001 ', ' 0001 ', ' 00001 ', ' 05 ', ' 06 ', ' 00025 ', ' 10 ', ' 20 ', ' 40 ', ' 05 '
]


def clear_blanks(word: str):
    for space in range(2, 10):
        word = word.replace(' ' * space, ' ')
    return word


def map_isins():
    umsatz_caceis_zusammenfassung = construct_file_path('Umsätze Caceis', 'Zusammenfassung.xlsm')
    isin_name_verzeichnis = construct_file_path('Bestandsdatenbank', 'ISIN_Name_Verzeichnis.xlsx')

    df1 = pd.read_excel(umsatz_caceis_zusammenfassung, sheet_name=0)
    df2 = pd.read_excel(isin_name_verzeichnis, sheet_name=0)

    # load workbook
    wb = px.load_workbook(isin_name_verzeichnis)
    ws1 = wb['Tabelle1']
    ws2 = wb['Tabelle2']
    ws3 = wb['CUSIP']

    # filter for the date < 7 days, remove duplicates and check, if isin is already in dataframe 2
    last_five_days = pd.Timestamp('now').floor('D') + pd.Timedelta(-7, unit='D')
    df1 = df1[df1['Valutadatum'] > last_five_days]
    df1 = df1.drop_duplicates(subset=['ISIN'])
    df1 = df1[~df1['ISIN'].isin(df2['ISIN'])]

    # Create three lists from dataframe and zip them to a tuple
    isin, security_name, currency = df1['ISIN'].to_list(), df1["Wertpapierbezeichnung"].to_list(), df1["Waehrung Kurs"].to_list()
    isin_securities_currency: zip = zip(isin, security_name, currency)

    for key, value, curr in isin_securities_currency:
        value = clear_blanks(value)

        for word in PROHIBITED_WORDS:
            value = value.replace(word, '')
            value = clear_blanks(value)

        value = clear_blanks(value)

        # Write in Table 1
        max_row1 = ws1.max_row + 1
        ws1.cell(max_row1, 1).value = key
        ws1.cell(max_row1, 2).value = value

        # Write in Table 2
        max_row2 = ws2.max_row + 1
        ws2.cell(max_row2, 1).value = key
        ws2.cell(max_row2, 2).value = value

        # if Cusip: Write in CUSIP
        if key[:2] == 'CA' or key[:2] == 'US' or curr == 'USD' or curr == 'CAD':
            cusip = key[2:-1]
            max_row3 = ws3.max_row + 1
            ws3.cell(max_row3, 1).value = cusip
            ws3.cell(max_row3, 2).value = key
            ws3.cell(max_row3, 3).value = value

        logging.info(f'Writing [{key}], [{value}] and [{curr}]')

    # save Workbook
    wb.save(isin_name_verzeichnis)
    logging.info(f'Saved workbook to path: {isin_name_verzeichnis}')
