import datetime
import logging
import requests

from openpyxl.reader.excel import load_workbook
from python.services.functions.mounted_directory import construct_file_path

apikey = {'X-CoinAPI-Key': 'E904D6C7-5201-462D-AC81-252F131E7925'}
CRYPTO_CURRENCIES = ['BTC', 'ETH', 'XRP', 'USDC', 'USDT', 'FLOW']
WORKSHEETS = ['EUR', 'USD']


def get_daily_crypto_currencies_exchange_rates():
    file_name = 'Crypto_Currencies_Exchange_Rates.xlsx'
    path_to_file = construct_file_path("1 Konvertierer", file_name)
    workbook = load_workbook(path_to_file)

    date_en: datetime.date = datetime.date.today() - datetime.timedelta(days=1)
    date_de = date_en.strftime("%d.%m.%Y")

    worksheet = workbook['EUR']
    rows: int = worksheet.max_row
    columns: int = len(CRYPTO_CURRENCIES)

    # check for all existing dates
    if date_de in [worksheet.cell(row=i, column=1).value for i in range(2, rows + 1)]:
        logging.warning(f'Exchange rate for date [{date_de}] was already inserted into the file: [{file_name}]')
        return None

    # for-loop für alle Währungen, die auch Workbooks sind
    for i in range(len(WORKSHEETS)):
        worksheet = workbook[WORKSHEETS[i]]
        worksheet.insert_rows(2)
        worksheet['A2'].value = date_de

        # for-loop für CRYPTO_CURRENCIES
        for j in range(2, columns + 2):
            url = rf'https://rest.coinapi.io/v1/ohlcv/KRAKEN_SPOT_{CRYPTO_CURRENCIES[j - 2]}_' \
                  rf'{WORKSHEETS[i]}/history?period_id=1DAY&time_start={date_en}T00:00:00'
            logging.info(f'Requesting data for {url}')

            response = requests.get(url, headers=apikey)
            result = response.json()

            for k in range(2, columns + 2):
                if worksheet.cell(row=1, column=k).value == CRYPTO_CURRENCIES[j - 2]:
                    worksheet.cell(row=2, column=k).value = result[0]['price_close']
                    break

    workbook.save(path_to_file)
    logging.info(f"{date_de} was inserted into the file: [{file_name}]")
    return True
