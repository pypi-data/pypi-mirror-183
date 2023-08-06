from typing import Optional

import openpyxl

from maphis.common.common import Info
from maphis.common.plugin import GeneralAction, ActionContext
from maphis.common.state import State
from maphis.plugins.maphis.general.common import _ndarray_export_routine, _filter_by_NDArray, \
    _group_measurements_by_sheet, _tabular_export_routine, show_export_success_message, get_prop_tuple_list


class ExportXLSX(GeneralAction):
    """
    GROUP: export
    NAME: Export xlsx
    KEY: export_xlsx
    DESCRIPTION: Exports measurements in XLSX format.
    """

    def __init__(self, info: Optional[Info] = None):
        super().__init__(info)

    def __call__(self, state: State, context: ActionContext) -> None:
        wb = openpyxl.Workbook()
        prop_tuple_list = get_prop_tuple_list(context)
        # self._tabular_export_routine(ws.append)
        nd_props, other_props = _filter_by_NDArray(prop_tuple_list, context)

        sheet_grouped_props = _group_measurements_by_sheet(other_props, context)

        for sheet_name, prop_list in sheet_grouped_props.items():
            ws = wb.create_sheet(sheet_name)
            _tabular_export_routine(prop_list, ws.append, context)

        sheet_nd_props = _group_measurements_by_sheet(nd_props, context)

        for sheet_name, prop_list in sheet_nd_props.items():
            ws = wb.create_sheet(sheet_name)
            _ndarray_export_routine(prop_list, ws.append, context)

        if 'Sheet' in wb:
            wb.remove(wb['Sheet'])

        wb.save(str(context.storage.location / f'{context.current_label_name}_results.xlsx'))

        show_export_success_message(context.storage.location, [f'{context.current_label_name}_results.xlsx'], context)

    # @property
    # def group(self) -> str:
    #     return 'export'