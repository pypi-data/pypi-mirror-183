import dataclasses
import re
import typing
from pathlib import Path
from typing import Optional
import importlib.resources

import numpy as np
import openpyxl
import pyparsing
from PySide6.QtCore import QUrl
from PySide6.QtWidgets import QMessageBox, QWidget, QPushButton, QVBoxLayout, QDialogButtonBox, \
    QDialog, QApplication, QScrollArea, QSizePolicy, QHBoxLayout, QFileDialog

from maphis.common.common import Info
from maphis.common.photo import Photo
from maphis.common.plugin import GeneralAction, ActionContext
from maphis.common.state import State
from maphis.common.storage import Storage
from maphis.common.utils import open_with_default_app
from maphis.plugins.profile_register.general.profiles import get_median_profile, merge_profiles
from maphis.plugins.profile_register.general.tag_mapping_widget import TagMappingWidget
from maphis.tags.tag_chooser import Tags
from maphis.tags.tags_choosers import GroupsOfTags


BODY = 16842752
PROFILE_KEY = "maphis.plugins.profile_register.properties.profile.profile"
TAG_REGEX = '\w+\D+\d*'
TAG_LIST_REGEX = f'\[({TAG_REGEX}\s*,\s*)*({TAG_REGEX})\]'
REGISTER_CONF_LINE_REGEX = re.compile(f'^({TAG_LIST_REGEX})\s*(\[({TAG_LIST_REGEX}\s*,\s*)*({TAG_LIST_REGEX})\])')

BR_L = pyparsing.Suppress(pyparsing.Literal('['))
BR_R = pyparsing.Suppress(pyparsing.Literal(']'))
TAG = pyparsing.Word(pyparsing.alphas, pyparsing.alphanums + '_')
TAG_LIST = pyparsing.Group(BR_L + pyparsing.delimited_list(TAG, delim=',') + BR_R)
CONF_LINE = TAG_LIST + BR_L + pyparsing.Group(pyparsing.delimited_list(TAG_LIST, delim=',')) + BR_R


@dataclasses.dataclass
class SettingWidget:
    toplevel_dialog: typing.Optional[QDialog]
    scroll_area: typing.Optional[QScrollArea]
    main_widget: typing.Optional[QWidget]
    main_layout: typing.Optional[QVBoxLayout]
    tag_mapping_layout: typing.Optional[QVBoxLayout]
    btnAdd: typing.Optional[QPushButton]
    btnSaveConf: typing.Optional[QPushButton]
    btnLoadConf: typing.Optional[QPushButton]
    diagButtonBox: typing.Optional[QDialogButtonBox]


class ProfileFusion(GeneralAction):
    """
    NAME: Profile fusion
    DESCRIPTION: Fuse body profiles based on their tags. A median profile is created for each group of images that
    matches the given tags. (how to specify and use the tags must be found)
    KEY: profile_fusion
    """
    def __init__(self, info: Optional[Info] = None):
        super().__init__(info)
        # self._setting_widget: QDialog = self._build_settings_widget()
        self._setting_widget: typing.Optional[SettingWidget] = None
        self._state: typing.Optional[State] = None

    def __call__(self, state: State, _: ActionContext):
        if not self._check_for_profile_measurement(state):
            return
        self._state = state
        self._execute(state, _)

    def _get_median_profile_for(self, iteration_tag: str, group_tags: typing.Set[str], state: State) -> typing.Optional[np.ndarray]:
        tags_set = set(group_tags).union({iteration_tag})

        photos = state.storage.photos_satisfying_tags(tags_set)

        if len(photos) == 0:
            return None
        profiles: np.ndarray = np.zeros((len(photos), 40), np.float32)
        print(f'number of photos satisfying the tag set {tags_set} = {len(photos)}')
        for i, photo in enumerate(photos):
            profiles[i] = np.array(photo['Labels'].region_props[BODY][PROFILE_KEY].value[0])
        median = get_median_profile(profiles, show_fig=False)
        return median

    def _check_for_profile_measurement(self, state: State) -> bool:
        photos_missing_profiles: typing.List[Photo] = []
        for photo in state.storage.images:
            if BODY not in photo['Labels'].region_props or PROFILE_KEY not in photo['Labels'].region_props[BODY]:
                photos_missing_profiles.append(photo)
        if len(photos_missing_profiles) > 0:
            QMessageBox.information(None, 'Missing profile measurements', f'{len(photos_missing_profiles)} photo(s) do(es) not have body profile measurement. Please compute them first.')
            return False
        return True

    def _execute(self, state: State, _: ActionContext):
        self._setting_widget = self._build_settings_widget2(state, state.storage)
        if self._setting_widget.toplevel_dialog.exec_() == QDialog.DialogCode.Accepted:
            output_folder = Path(state.storage.location) / 'registered_profiles'
            if not output_folder.exists():
                output_folder.mkdir(exist_ok=True)
            # TODO add an option to overwrite the existing results
            # if self.chk_delete_existing.isChecked():
            #     os.remove(output_folder / 'profiles.xlsx')
            if (worksheet_path := output_folder / 'profiles.xlsx').exists():
                wb = openpyxl.load_workbook(worksheet_path, read_only=False)
            else:
                wb = openpyxl.Workbook()
                ws_med_prof = wb.active
                ws_med_prof.title = 'Median profiles'
                ws_med_prof.append(['ProfileID'] + [f'G_BP_{i}' for i in range(40)])

                ws_aligned = wb.create_sheet('Aligned profiles')
                ws_aligned.append(['ProfileID', 'AlignedTo'] + [f'G_BP_{i}' for i in range(40)])

            ws_med_prof = wb['Median profiles']
            ws_aligned = wb['Aligned profiles']

            mappings = self._gather_mappings()
            for mapping in mappings:
                src, dsts = mapping
                src_median, dst_median, aligned = self._register(src, dsts, state)
                src_median_profile_id = ','.join(src)
                dst_median_profile_id = '_'.join([','.join(dst_tags) for dst_tags in dsts])
                ws_med_prof.append([src_median_profile_id] + [str(float(val)) for val in src_median])
                ws_med_prof.append([dst_median_profile_id] + [str(float(val)) for val in dst_median])
                ws_aligned.append([src_median_profile_id, dst_median_profile_id] + [str(float(val)) for val in aligned])
            wb.save(worksheet_path)
            wb.close()
            self._show_success_info_dialog(worksheet_path)
        else:
            pass

    def _gather_mappings(self) -> typing.List[typing.Tuple[Tags, GroupsOfTags]]:
        mappings: typing.List[typing.Tuple[Tags, GroupsOfTags]] = []
        for i in range(self._setting_widget.tag_mapping_layout.count()):
            tag_mapping: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(i).widget()
            mappings.append(tag_mapping.get_mapping())
        return mappings

    def _register(self, src: typing.List[str], dsts: typing.List[typing.List[str]], state: State):
        src_profiles = self._gather_profiles(src, state)
        src_median_profile = get_median_profile(src_profiles)

        dst_median_profiles: typing.List[np.ndarray] = []
        for dst in dsts:
            dst_profiles = self._gather_profiles(dst, state)
            dst_median_profiles.append(get_median_profile(dst_profiles))
        dst_median_profile = get_median_profile(np.array(dst_median_profiles))

        aligned = merge_profiles(src_median_profile, dst_median_profile, x_weight=0.5, y_weight=0.5)

        return src_median_profile, dst_median_profile, aligned

    def _gather_profiles(self, tags: typing.List[str], state: State) -> np.ndarray:
        photos: typing.List[Photo] = state.storage.photos_satisfying_tags(set(tags))
        profiles: typing.List[np.ndarray] = [np.array(photo['Labels'].region_props[BODY][PROFILE_KEY].value[0]) for photo in photos]
        return np.array(profiles)

    def _show_success_info_dialog(self, output_path: Path):
        diag_box = QMessageBox(QMessageBox.Icon.Information, "Profile fusion completed",
                               f"The results are stored in {output_path}. Do you wish to open the folder?",
                               QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QApplication.activeWindow())
        diag_box.button(QMessageBox.StandardButton.Yes).clicked.connect(diag_box.accept)
        diag_box.button(QMessageBox.StandardButton.No).clicked.connect(diag_box.reject)

        if diag_box.exec_() == QMessageBox.DialogCode.Accepted:
            open_with_default_app(output_path.parent)
        diag_box.close()
        diag_box.deleteLater()

    def setting_widget(self) -> typing.Optional[QWidget]:
        return None

    def _remove_tag_mapping(self, tag_map_widget: TagMappingWidget):
        if self._setting_widget.tag_mapping_layout.count() == 1:
            return
        tag_map_widget.hide()
        self._setting_widget.tag_mapping_layout.removeWidget(tag_map_widget)
        tag_map_widget.deleteLater()
        last_mapping: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(self._setting_widget.tag_mapping_layout.count() - 1).widget()
        last_mapping.enable_remove_button(self._setting_widget.tag_mapping_layout.count() > 1)
        self._update_apply_button()

    def _add_tag_mapping(self, state: State, storage: Storage):
        new_mapping = TagMappingWidget(state, parent=self._setting_widget.main_widget)
        new_mapping.remove_this.connect(self._remove_tag_mapping)
        new_mapping.complete_status_changed.connect(lambda _: self._update_apply_button())
        self._setting_widget.tag_mapping_layout.addWidget(new_mapping)

        for idx in range(self._setting_widget.tag_mapping_layout.count()):
            tag_map_widget: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(idx).widget()
            tag_map_widget.enable_remove_button(self._setting_widget.tag_mapping_layout.count() > 1)
        self._update_apply_button()

    def _update_apply_button(self):
        all_mappings_complete = True
        for i in range(self._setting_widget.tag_mapping_layout.count()):
            map_widget: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(i).widget()
            all_mappings_complete = all_mappings_complete and map_widget.is_complete_and_matches_photos
        self._setting_widget.diagButtonBox.button(QDialogButtonBox.StandardButton.Apply).setEnabled(all_mappings_complete)
        self._setting_widget.btnSaveConf.setEnabled(all_mappings_complete)

    def _build_settings_widget2(self, state: State, storage: Storage) -> SettingWidget:
        sett_widget = SettingWidget(
            toplevel_dialog=QDialog(),
            scroll_area=QScrollArea(),
            main_widget=QWidget(),
            main_layout=QVBoxLayout(),
            tag_mapping_layout=QVBoxLayout(),
            btnAdd=QPushButton('+'),
            btnSaveConf=QPushButton('Save configuration...'),
            btnLoadConf=QPushButton('Load configuration...'),
            diagButtonBox=QDialogButtonBox(QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Cancel)
        )
        sett_widget.toplevel_dialog.setLayout(QVBoxLayout())
        sett_widget.toplevel_dialog.layout().addWidget(sett_widget.scroll_area)
        sett_widget.toplevel_dialog.setMinimumHeight(600)
        sett_widget.toplevel_dialog.setMinimumWidth(800)

        sett_widget.main_widget.setSizePolicy(QSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding))
        sett_widget.scroll_area.setSizeAdjustPolicy(QScrollArea.SizeAdjustPolicy.AdjustToContents)
        sett_widget.scroll_area.setWidgetResizable(True)

        sett_widget.main_widget.setLayout(sett_widget.main_layout)

        sett_widget.main_layout.addLayout(sett_widget.tag_mapping_layout)
        sett_widget.main_layout.addWidget(sett_widget.btnAdd)
        sett_widget.main_layout.addStretch()

        sett_widget.btnAdd.clicked.connect(lambda: self._add_tag_mapping(state, storage))

        sett_widget.scroll_area.setWidget(sett_widget.main_widget)

        top_level_buttons_layout = QHBoxLayout()
        sett_widget.toplevel_dialog.layout().addLayout(top_level_buttons_layout)

        sett_widget.btnSaveConf.clicked.connect(self._save_configuration)
        sett_widget.btnSaveConf.setEnabled(False)

        sett_widget.btnLoadConf.clicked.connect(self._load_configuration)

        top_level_buttons_layout.addWidget(sett_widget.btnSaveConf)
        top_level_buttons_layout.addWidget(sett_widget.btnLoadConf)

        top_level_buttons_layout.addStretch()

        button_diag_box = sett_widget.diagButtonBox #QDialogButtonBox(QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Cancel)
        top_level_buttons_layout.addWidget(button_diag_box)
        button_diag_box.button(QDialogButtonBox.StandardButton.Apply).clicked.connect(sett_widget.toplevel_dialog.accept)
        button_diag_box.button(QDialogButtonBox.StandardButton.Apply).setEnabled(False)
        button_diag_box.button(QDialogButtonBox.StandardButton.Cancel).clicked.connect(sett_widget.toplevel_dialog.reject)

        # self._setting_widget.diagButtonBox = button_diag_box

        self._setting_widget = sett_widget

        self._add_tag_mapping(state, storage)

        return sett_widget

    def _load_configuration(self):
        if (recent_folder := self._get_recent_open_save_location()) is None:
            recent_folder = QUrl()
        else:
            recent_folder = QUrl.fromLocalFile(str(recent_folder))
        result = QFileDialog.getOpenFileUrl(self._setting_widget.toplevel_dialog, 'Load configuration from...', dir=recent_folder)
        qurl: QUrl = result[0]
        if qurl.isEmpty():
            return
        load_path = Path(qurl.toLocalFile())
        self._save_recent_open_save_location(load_path.parent)

        with open(load_path, 'r') as f:
            str_mappings = [line.strip() for line in f.readlines()]
        valid_mappings: typing.List[typing.Tuple[Tags, GroupsOfTags]] = []
        invalid_mappings: typing.List[str] = []
        for mapping in str_mappings:
            if not CONF_LINE.matches(mapping):
                invalid_mappings.append(mapping)
            else:
                valid_mappings.append(tuple(CONF_LINE.parse_string(mapping).as_list()))

        if len(invalid_mappings) > 0:
            # TODO show a message box
            return
        how_many_mapping_widgets_to_add = len(valid_mappings) - self._setting_widget.tag_mapping_layout.count()
        for i in range(how_many_mapping_widgets_to_add):
            self._add_tag_mapping(self._state, None)

        for i, mapping in enumerate(valid_mappings):
            tag_map_widg: TagMappingWidget = self._setting_widget.tag_mapping_layout.itemAt(i).widget()
            tag_map_widg.set_mapping(mapping)

    def _save_configuration(self):
        mappings = self._gather_mappings()
        if (recent_folder := self._get_recent_open_save_location()) is None:
            recent_folder = QUrl()
        else:
            recent_folder = QUrl.fromLocalFile(str(recent_folder))
        result = QFileDialog.getSaveFileUrl(self._setting_widget.toplevel_dialog, 'Save configuration as...', dir=recent_folder)
        qurl: QUrl = result[0]
        if qurl.isEmpty():
            return
        save_path = Path(qurl.toLocalFile())
        self._save_recent_open_save_location(save_path.parent)

        with open(save_path, 'w') as f:
            for tags, groups_of_tags in mappings:
                line = f'{dequote(str(tags))} {dequote(str(groups_of_tags))}\n'
                f.write(line)

    def _get_recent_open_save_location(self) -> typing.Optional[Path]:
        with importlib.resources.path('maphis.plugins.profile_register.general', '.') as path:
            if not (recent_path_file := path / 'recent_save_open_path.txt').exists():
                return Path.home()
            with open(recent_path_file, 'r') as f:
                recent_path = Path(f.read())
            return recent_path

    def _save_recent_open_save_location(self, loc: Path):
        with importlib.resources.path('maphis.plugins.profile_register.general', '.') as path:
            with open(path / 'recent_save_open_path.txt', 'w') as f:
                f.write(str(loc))


def dequote(text: str) -> str:
    return text.replace('"', '').replace("'", '')