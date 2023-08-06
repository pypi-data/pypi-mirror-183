from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
import xlwings as xw
import logging
from config import *

logging.basicConfig(format=FORMAT, level=logging.DEBUG, filename=LOG_PATH)

def openpyxl_get_datas_workbook_worksheet(filename, sheet_name=None, data_only=True, read_only=True):
    """ 获取Excel数据和WorkSheet WorkBook

        Args:
            filename [str]: Excel文件名
            sheet_name [str]: 工作表名, 如果未传入, 默认读取第一个工作表
            data_only [bool]: 是否仅将数据转化为公式, True代表公式会进行计算
            read_only [bool]: 是否仅可读, 如果为True, 读取数据速度会提升很多,但不可以进行清除worksheet数据等操作
        
        Returns:
            worksheet_datas [list[list, list, ...]]: 工作表数据
            workbook [Workbook]: 工作薄
            worksheet [Worksheet]: 工作表
        
    """
    workbook = load_workbook(filename, read_only=read_only, data_only=data_only)

    if sheet_name is None:
        worksheet = workbook[workbook.sheetnames[0]]
    else:
        worksheet = workbook[sheet_name]

    worksheet_values = worksheet.values
    worksheet_datas = [[x for x in y] for y in worksheet_values]
    
    return worksheet_datas, workbook, worksheet

def openpyxl_get_datas(worksheet: Worksheet) -> list:
    """ 获取工作表数据

        Args:
            worksheet [Worksheet]: 需要读取数据的工作表
        
        Returns:
            worksheet_datas [list[list, list, ...]]: 工作表数据
    """
    worksheet_values = worksheet.values
    worksheet_datas = [[x for x in y] for y in worksheet_values]
    
    return worksheet_datas

def openpyxl_replace_worksheet_data(origin_worksheet: Worksheet, to_be_replaced_worksheet: Worksheet):
    """ 读取原始工作表的数据并替换另一个工作表的数据,首先会清空被替换工作表的数据 基于openpyxl

        Args:
            origin_worksheet [xw.Sheet]: 原始工作表
            to_be_replaced_worksheet [xw.Sheet]: 被替换工作表

    """
    openpyxl_clean_worksheet_datas(to_be_replaced_worksheet)
    datas = openpyxl_get_datas(origin_worksheet)
    openpyxl_datas_to_worksheet(datas=datas, worksheet=to_be_replaced_worksheet)

def openpyxl_clean_worksheet_datas(worksheet: Worksheet):
    """ 删除工作表中的所有数据, 因为使用delete_rows函数, 不会打乱原有数据基础上的公式

        Args:
            worksheet [Worksheet]: 要清除所有数据的工作表
    """
    max_row = worksheet.max_row
    worksheet.delete_rows(1, max_row)

def openpyxl_datas_to_worksheet(datas: list, worksheet: Worksheet):
    """ 将读取出来的数据列表填入到工作表中 如果有合并单元格 数据可能与原格式不同

        Args:
            datas list[list, list, ...]: 要填入的数据
            worksheet [Worksheet]: 被数据填入的工作表
    """
    for row in range(len(datas)):
        for col in range(len(datas[row])):
            worksheet.cell(row=row + 1, column=col + 1).value = datas[row][col]

def openpyxl_parser_merged_cell(worksheet: Worksheet, row: int, col: int):
    """ 获取合并单元格的数据 基于openpyxl 需要load_workbook(read_only=False)

        Example:
                     |  A  |  B  |
              1      |    test   |
              2      |     ^     |      ^ 表示上下均为test(合并单元格)
              3      |  1  |  2  |

            openpyxl_parser_merged_cell(worksheet, 1, 1)  -> test
            openpyxl_parser_merged_cell(worksheet, 1, 2)  -> test
            openpyxl_parser_merged_cell(worksheet, 2, 1)  -> test
            openpyxl_parser_merged_cell(worksheet, 2, 2)  -> test
            openpyxl_parser_merged_cell(worksheet, 3, 1)  -> 1
            openpyxl_parser_merged_cell(worksheet, 3, 2)  -> 1

        Args:
            worksheet [Worksheet]: 工作表
            row [int]: 行
            col [int]: 列

        Returns:
            cell.value []: 单元格的值, 如果为合并单元格, 则一直查找到左上角单元格的值
    """
    cell = worksheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in worksheet.merged_cell_ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value

def openpyxl_parse_multi_title_to_single(worksheet: Worksheet, titleList: list, nrows: int, title_start_row: int):
    """ 将多行标题更改为一行后导入pandas操作 基于openpyxl

    Example:
            before:   |    test   |
                      |  1  |  2  |

             after:   ["test_1", "test_2"]
            
        Args:
            worksheet [Worksheet]: 多行标题所在工作表
            titleList [list]: 标题列表
            nrows [int]: 标题共占几行
            title_start_row [int]: 标题所在工作表行(标题第一行处)
        Returns:
            single_list [list]: 单行标题列表
    """
    fillList = []
    tmpList = []
    for i in range(len(titleList) - 1):
        tmpList = []
        for j in range(len(titleList[i])):
            tmpList.append(openpyxl_parser_merged_cell(worksheet, title_start_row + nrows + 1 + i, j + 1))
            # print(sheet.cell(SearchAssembly + 2 + 1 + i, j + 1).value)
        fillList.append(tmpList)
    fillList.extend(titleList[len(titleList) - 1:])
    
    text = ""
    single_list = []
    
    for i in range(len(fillList[0])):
        single_list.append("")
    
    for i in range(len(fillList)):
        for j in range(len(fillList[i])):
            # print(fillList[i][j])
            if fillList[i][j] != None:
                text = single_list[j].strip()
                text += f"_{fillList[i][j]}"
                single_list[j] = text.strip().strip('_')
                
    return single_list

def xlwings_get_app_workbook_worksheet(filename: str, sheet_name=None, visible=False, add_book=False, display_alerts=False, screen_updating=False):
    """ 使用xlwings获取`xw.App` `xw.Book` `xw.Sheet`
        目前仅支持Windows Mac
    
        Args:
            filename [str]: 需要读取的Excel文件名
            sheet_name [str]: 默认为None, 获取工作表名
            visible [bool]: 见`xw.App`解释
            add_book [bool]: 见`xw.App`解释
            display_alerts [bool]: 见`xw.App`解释
            screen_updating [bool]: 见`xw.App`解释

        Returns:
            app [xw.App]: 见`xw.App`解释
            workbook [xw.Book]: 见`xw.Book`解释
            worksheet [xw.Sheet]: 见`xw.Sheet`解释
    """
    app = xw.App(visible=visible, add_book=add_book)
    # with xw.App(visible=visible, add_book=add_book) as app:
    app.display_alerts = display_alerts
    app.screen_updating = screen_updating

    if LANGUAGE == 'cn':
        logging.debug("使用xlwings获取xw.App xw.Book xw.Sheet")
        logging.debug(f"文件名: {filename}")
        logging.debug(f"工作表名: {sheet_name}")
    elif LANGUAGE == 'en':
        logging.debug("Use xlwings get xw.App xw.Book xw.Sheet")
        logging.debug(f"Filename: {filename}")
        logging.debug(f"Sheet name: {sheet_name}")

    logging.debug(f"visible: {visible}")
    logging.debug(f"add_book: {add_book}")
    logging.debug(f"display_alerts: {display_alerts}")
    logging.debug(f"screen_updating: {screen_updating}")

    workbook = app.books.open(filename)

    if sheet_name is None:
        worksheet = workbook.sheets[0]
    else:
        worksheet = workbook.sheets[sheet_name]

    return app, workbook, worksheet

def xlwings_replace_worksheet_data(origin_worksheet: xw.Sheet, to_be_replaced_worksheet: xw.Sheet, clear_rule='clear_contents'):
    """ 读取原始工作表的数据并替换另一个工作表的数据,首先会清空被替换工作表的数据 基于xlwings

        Args:
            origin_worksheet [xw.Sheet]: 原始工作表
            to_be_replaced_worksheet [xw.Sheet]: 被替换工作表
            clear_rule [str]: 清空规则, 可选`clear|clear_contents|clear_formats` 分别为清空格式与数据 仅清空数据 仅清空格式
    
    """
    # origin_worksheet_max_row = origin_worksheet.used_range.last_cell.row
    # origin_worksheet_max_col = origin_worksheet.used_range.last_cell.col
    if LANGUAGE == 'cn':
        logging.debug("替换工作表中的数据")
        logging.debug(f"清除原工作表规则为{clear_rule}")
    elif LANGUAGE == 'en':
        logging.debug("Replace data in worksheet")
        logging.debug(f"Clear the original worksheet rule as {clear_rule}")
    if clear_rule == 'clear':
        to_be_replaced_worksheet.clear()
    elif clear_rule == 'clear_contents':
        to_be_replaced_worksheet.clear_contents()
    elif clear_rule == 'clear_formats':
        to_be_replaced_worksheet.clear_formats()
    else:
        to_be_replaced_worksheet.clear_contents()

    to_be_replaced_worksheet.range('A1').value = origin_worksheet.used_range.value

def xlwings_sort(worksheet: xw.Sheet, sort_range: str, sort_col: str, order='descending'):
    """ 工作表排序

        Args:
            worksheet [xw.Sheet]: 需要进行排序的工作表
            sort_range [str]: 排序的范围 Example: 'A1:L11'
            sort_col [str]: 对哪个列进行排序 Example: 'C1:C11'
            order [str]: default=descending 升降序 默认为降序 ascending|descending 判断时只要不为ascending即为降序
    """
    if LANGUAGE == 'cn':
        logging.debug(f"对{sort_range}中的数据排序")
        if order == 'ascending':
            logging.debug(f"根据{sort_col}升序排序")
        else:
            logging.debug(f"根据{sort_col}降序排序")
    elif LANGUAGE == 'en':
        logging.debug(f"Sort the data in {sort_range}")
        if order == 'ascending':
            logging.debug(f"Sort in ascending order according to {sort_col}")
        else:
            logging.debug(f"Sort in descending order according to {sort_col}")
            
    if order == "ascending":
        worksheet.range(sort_range).api.Sort(Key1=worksheet.range(sort_col).api, Order1=1, Orientation=1)
    else:
        worksheet.range(sort_range).api.Sort(Key1=worksheet.range(sort_col).api, Order1=2, Orientation=1)


if __name__ == "__main__":
    # datas, workbook, worksheet = openpyxl_get_datas_workbook_worksheet('./收入积分_20221212-20221218.xlsx', data_only=False, read_only=True)
    # new_datas, new_workbook, new_worksheet = openpyxl_get_datas_workbook_worksheet('./积分.xlsx', sheet_name="12月", data_only=False, read_only=False)
    # openpyxl_replace_worksheet_data(worksheet, new_worksheet)
    # new_workbook.save('./test.xlsx')
    # origin_app, origin_workbook, origin_worksheet = xlwings_get_app_workbook_worksheet('./收入积分_copy_基础表_汇总数据_20221222231901.xlsx')
    to_be_replaced_app, to_be_replaced_workbook, to_be_replaced_worksheet = xlwings_get_app_workbook_worksheet('./骏马奖累计.xlsx', sheet_name='排序')
    # xlwings_replace_worksheet_data(origin_worksheet, to_be_replaced_worksheet)
    xlwings_sort(to_be_replaced_worksheet, "A4:L11", "L4:L11", 'descending')
    to_be_replaced_workbook.save()

    to_be_replaced_workbook.close()
    # origin_workbook.close()

    # origin_app.quit()
    to_be_replaced_app.quit()