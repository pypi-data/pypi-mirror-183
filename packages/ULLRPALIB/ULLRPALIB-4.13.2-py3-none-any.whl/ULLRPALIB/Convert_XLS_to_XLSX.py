from openpyxl import load_workbook
import pyexcel as p
# pip install pyexcel pyexcel-xls pyexcel-xlsx
import sys

def unifica_hojas(archivo, desplazamiento):
    p.save_book_as(file_name = archivo, dest_file_name = archivo+"x")   # Ya elimina la pestaña de Macro

    book = load_workbook(filename=archivo+"x")
    hojas = book.sheetnames
    print(hojas)
    destino = book[hojas[0]]
    if len(hojas) > 1:   # Hay mas de una hoja
        for org in hojas[1:]:
            origen = book[org]
            print("De ", org, " a ", hojas[0], " en ", destino.max_row + 1)
            fila_destino = destino.max_row
            for fila in origen.iter_rows(min_row = desplazamiento):
                fila_destino += 1
                for cell in fila:
                    destino.cell(row = fila_destino, column = cell.column).value = cell.value
            book.remove(origen)

    if desplazamiento == 4:  # Hay cabecera de búsqueda
        print("Eliminando cabecera de búsqueda")
        destino.delete_rows(1, amount=2)
    book.save(archivo+"x")

'''
if len(sys.argv) != 3:
    print("Error: usage Convert_XLS_to_XLSX file.xls eliminar_busqueda|no_eliminar_busqueda")
    unifica_hojas("c:\\local\\Sede_Docuconta_Tareas_Una_Hoja.xls", 2)
else:
    fichero = sys.arv[1]
    busqueda = sys.argv[2]
    if busqueda == "eliminar_busqueda":
        unifica_hojas(fichero, 4)
    else:
        unifica_hojas(fichero, 2)
'''