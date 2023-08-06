from tkinter import E
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


VERBOSE = False

def limpia_hoja_excel(hoja):
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.value = ""


def leer_excel(nombre_archivo, nombre_hoja , campos_clave: list):
    book = load_workbook(filename=nombre_archivo, read_only=True)
    hoja = book[nombre_hoja]
    # print('dimensiones:', hoja.calculate_dimension())
    titulos = hoja[1]
    columnas_clave = []
    dict = {}
    for titulo in titulos:
        for campo in campos_clave:
            if titulo.value == campo:
                # print('campo encontrado:', titulo.column_letter)
                columnas_clave.append(titulo.column_letter)

    # print(columnas_clave)
    if len(columnas_clave) != len(campos_clave):
        print('ERROR: No se ha podido encontrar alguno de los campos clave en el Excel.\nSe interrumpe la lectura.')
        return

    # dict['titulos'] = [i.value for i in titulos]

    num_claves_repetidas = 0
    for fila in hoja.iter_rows(min_row=2):
        indice_titulos = 0
        clave = ''
        dict_valores = {}
        for celda in fila:
            # En el caso de que haya celdas sin valor de columna
            if(celda.value is None):  # celda.data_type == 'n'
                #print('AAAAAAAAA celda vacía:', celda.value, fila)
                dict_valores[titulos[indice_titulos].value] = ''
                indice_titulos += 1
                continue
            elif celda.column_letter in columnas_clave:
                # print('celda clave', celda.value)
                if isinstance(celda.value, datetime.datetime):
                    clave = clave + \
                        celda.value.strftime('%d/%m/%Y %H:%M:%S') + ';'
                else:
                    clave = clave + str(celda.value) + ';'

            dict_valores[titulos[indice_titulos].value] = celda.value
            indice_titulos += 1
            # print(celda.coordinate)
        clave = clave[:-1]
        #print('clave:', clave)
        if clave not in dict:
            dict[clave] = dict_valores
        else:
            if VERBOSE:
                print('AVISO: clave ya existente en el diccionario, no se añade la fila:',
                      clave, '-', dict_valores)
            num_claves_repetidas += 1
    book.close()

    lista_vacios = []
    for clave in dict:
        if clave == '':
            lista_vacios.append(clave)

    for item in lista_vacios:
        dict.pop(item)
    # Se calcula el numero total de filas leídas sumando el numero de claves en el diccionario,
    # el numero de filas no añadidas debido a repetición de clave
    # y una más que representa la fila de títulos
    if VERBOSE:
        if (list(dict.values())):
            print(f'\nSe ha leído la hoja "{nombre_hoja}" en el fichero excel "{nombre_archivo}", usando como claves: {campos_clave}\n' +
                f'Se han leído {len(dict.keys()) + num_claves_repetidas + 1} filas y {len(list(dict.values())[0])} columnas\n' +
                f'Hay {len(dict.keys())} claves únicas, y se omitieron {num_claves_repetidas} filas debido a repetición de clave\n')
    return dict

